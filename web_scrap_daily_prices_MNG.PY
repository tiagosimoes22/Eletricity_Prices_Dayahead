# Import necessary libraries
import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime,timedelta
from dateutil.relativedelta import relativedelta
import matplotlib.pyplot as plt
import numpy as np
import matplotlib.dates as mdates
import os
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl.utils import get_column_letter
import os
import pandas as pd
import re
import smtplib
from email.message import EmailMessage
import imghdr
import warnings
warnings.filterwarnings("ignore")
pd.set_option('display.max_columns', None)
pd.set_option('display.max_rows', None)

# Data de amanha tendo em conta a data de hoje
date_today = datetime.now().date()
date_tommorow = date_today + relativedelta(days=1)

# Web scraping setup
current_date = date_tommorow # Starting from tomorrow's date
df_all_vf = pd.DataFrame()  # Initialize an empty DataFrame to store all data

# Iniciar o navegador
driver = webdriver.Chrome()
driver.get("https://www.omie.es/pt/market-results/daily/daily-market/day-ahead-price?scope=daily&date={current_date}".format(current_date=current_date))

wait = WebDriverWait(driver, 10)

# Aceita cookies se necessário (dependendo do site)
try:
        accept_cookies = wait.until(EC.element_to_be_clickable((By.ID, "Aceitar")))
        accept_cookies.click()
except:
        pass

# Click no icon de download
svg_parent = wait.until(EC.element_to_be_clickable((
    By.CSS_SELECTOR, "g.highcharts-button"
)))
svg_parent.click()

# Clicar em "Ver a tabela correspondente"
ver_tabela = wait.until(EC.element_to_be_clickable((
    By.XPATH, "//div[contains(text(), 'Ver a tabela correspondente')]"
)))
ver_tabela.click()

tabela_elemento = wait.until(EC.presence_of_element_located((
    By.CSS_SELECTOR, "div.highcharts-data-table"
)))

# Extrair a tabela como HTML
tabela_html = tabela_elemento.get_attribute('outerHTML')

# Agora usa o pandas para transformar em DataFrame
df = pd.read_html(tabela_html)[0]

# Adicionar uma coluna com o dia do current_date
df.insert(0, 'Day', current_date)  # Insere a coluna 'Day' como a primeira coluna

# Concatenar os dados ao DataFrame principal
df_all = df


# Reformatação do DataFrame
df_all = df[['Day', 'Período', 'Preços marginais sistema português']]

df_all['Hora'] = df_all['Período'] -1

#df_all['Group_4h'] = 
df_all['Slots_4h'] = df_all['Hora'] //4
df_all['Slots_2h'] = df_all['Hora'] //2

# Criar os intervalos de 0 a 24 em blocos de 4 horas
bins = list(range(0, 25, 4))  # [0, 4, 8, 12, 16, 20, 24]
labels = [f'{bins[i]}h-{bins[i+1]}h' for i in range(len(bins)-1)]
bins_2 = list(range(0, 25, 2))  # [0, 4, 8, 12, 16, 20, 24]
labels_2 = [f'{bins_2[i]}h-{bins_2[i+1]}h' for i in range(len(bins_2)-1)]

# Create coluna 'Price_AVG' of the day
df_all['Price_AVG'] = df_all.groupby('Day')['Preços marginais sistema português'].transform('mean')

# Categorizar as horas nos slots
df_all['Slot_4h'] = pd.cut(df_all['Hora'], bins=bins, labels=labels, right=False)
df_all['Slot_2h'] = pd.cut(df_all['Hora'], bins=bins_2, labels=labels_2, right=False)

df_all['Preços_slots4'] = df_all.groupby('Slot_4h')['Preços marginais sistema português'].transform('mean')
df_all['Preços_slots2'] = df_all.groupby('Slot_2h')['Preços marginais sistema português'].transform('mean')

# Find the 4h slot with the highest average Portuguese marginal price
slot_maior_4h = df_all.groupby('Slot_4h')['Preços marginais sistema português'].mean().idxmax()
preco_maior_4h = df_all['Preços_slots4'].max().round(2)

# Find the 2h slot with the highest average Portuguese marginal price
slot_maior_2h = df_all.groupby('Slot_2h')['Preços marginais sistema português'].mean().idxmax()
preco_maior_2h = df_all['Preços_slots2'].max().round(2)

# Find the second highest 2h slot price
second_highest_value = df_all.loc[df_all['Preços_slots2'] == df_all['Preços_slots2'].nlargest(3).iloc[-1], 'Slot_2h'].iloc[0]
preco_maior_2h_segunda = df_all['Preços_slots2'].nlargest(3).iloc[-1].round(2)

# Get the slot of lowest price
df_all['low_avg_price'] = np.where(
	df_all['Preços marginais sistema português'] < df_all['Price_AVG'],
	"Yes",
	"No"
)

min_price_hours = f"{df_all.loc[df_all['low_avg_price'] == 'Yes', 'Hora'].min()}h-{df_all.loc[df_all['low_avg_price'] == 'Yes', 'Hora'].max()}h"

df_all_vf = df_all[['Day', 'Hora', 'Preços marginais sistema português']].drop_duplicates(subset=['Day', 'Hora']).sort_values(by=['Day', 'Hora'])

# Create the 'Hora' label as a string for each row
df_all_vf['Hora'] = df_all_vf['Hora'].apply(lambda h: f"{h}h-{h+1}h")

# Pivot the table
df_all_vf = df_all_vf.pivot(index='Day', columns='Hora', values='Preços marginais sistema português')

# Compute the daily average price
df_all_vf['Price_Daily_Avg'] = df_all_vf.iloc[:, 0:].mean(axis=1).round(2)
df_all_vf['Slot_4h_max'], df_all_vf['Slot_4h_price'], df_all_vf['Slot_2h_frist'], df_all_vf['Slot_2h_frist_price'], df_all_vf['Slot_2h_second'], df_all_vf['Slot_2h_second_price'], df_all_vf['Slot_min_price'] = slot_maior_4h,preco_maior_4h, slot_maior_2h, preco_maior_2h, second_highest_value, preco_maior_2h_segunda, min_price_hours 

# Sort only hour columns in order, then append the rest (summary/stat columns)
hour_pattern = re.compile(r'^\d+h-\d+h$')
hour_cols = [col for col in df_all_vf.columns if hour_pattern.match(col)]
hour_cols_sorted = sorted(hour_cols, key=lambda x: int(x.split('h')[0]))
other_cols = [col for col in df_all_vf.columns if col not in hour_cols]
df_all_vf = df_all_vf[hour_cols_sorted + other_cols]

file_path = "C://Users//tmlopes//OneDrive - Sonae Arauco//Spot_Electricity_Project//spot_price_dailyHour_PT.xlsx"
# Save DataFrame to Excel
df_all_vf.to_excel(file_path, index=True, sheet_name='Spot_PT')

# Load workbook and worksheet
book = load_workbook(file_path)
sheet = book['Spot_PT']

# Define the table range (A1 to last column/row)
max_row = sheet.max_row
max_col = sheet.max_column
table_ref = f"A1:{get_column_letter(max_col)}{max_row}"

# Create and add the table
table = Table(displayName="SpotPTTable", ref=table_ref)
style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=True)
table.tableStyleInfo = style
sheet.add_table(table)

# Save workbook
book.save(file_path)
book.close()       